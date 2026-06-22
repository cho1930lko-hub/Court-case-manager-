"""Excel data file (case_management.xlsx) ko initialize karta hai agar exist nahi karta."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_FILE = os.path.join(os.path.dirname(__file__), "data", "case_management.xlsx")

CASES_HEADERS = [
    "Case_ID", "Case_Type", "Satra_Pareeksha_Sankhya", "Mukadma_Apradh_Sankhya",
    "Rajya_Banam", "Dhara", "Thana", "Janpad", "Nyayalaya", "Tareekh_Peshi",
    "Status", "Notes"
]

WITNESS_HEADERS = [
    "Witness_ID", "Case_ID", "Naam", "Pata", "Mobile", "Bhumika", "Custom_Note"
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)


def style_header(sheet, headers):
    for idx, h in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(idx)].width = 22
    sheet.freeze_panes = "A2"


def init_excel():
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    if os.path.exists(DATA_FILE):
        return DATA_FILE

    wb = Workbook()
    cases_sheet = wb.active
    cases_sheet.title = "Cases"
    style_header(cases_sheet, CASES_HEADERS)

    witness_sheet = wb.create_sheet("Witnesses")
    style_header(witness_sheet, WITNESS_HEADERS)

    wb.save(DATA_FILE)
    return DATA_FILE


if __name__ == "__main__":
    path = init_excel()
    print(f"Initialized: {path}")
